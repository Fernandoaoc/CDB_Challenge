import logging
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelReader:

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)

    def read(self) -> list[dict]:

        logger.info("Reading Excel file %s", self.file_path)
        workbook = load_workbook(self.file_path)
        sheet: Worksheet = workbook.active

        headers: list[Any] = [
            cell.value
            for cell in sheet[1]
        ]

        data: list[dict[str, Any]] = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            record = {
                header: value
                for header, value in zip(headers, row)
            }

            data.append(record)

        workbook.close()
        logger.info("Excel file read complete. %d records loaded", len(data))

        return data
